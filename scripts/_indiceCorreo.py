# -*- coding: utf-8 -*-
"""
_indiceCorreo.py — arma el _INDICE.xlsx del archivo de correo a partir del _indice.jsonl
que deja `_archivoCorreo.py`.

POR QUE UN EXCEL
    El archivo de correo tiene que servir SIN Claude y SIN Outlook. El Excel es lo que Fak
    (o quien venga despues) abre para buscar por fecha, por quien escribio, por asunto o
    por nombre de adjunto, y la columna "archivo" dice exactamente cual .eml abrir.

    Fila 1 y columna A quedan VACIAS: los entregables de Barack empiezan en B2.

USO
    python scripts/_indiceCorreo.py --carpeta "<carpeta del archivo de correo>"
        --salida _INDICE.xlsx      (por defecto, dentro de la misma carpeta)

Corre con el Python del sistema (3.13): openpyxl esta ahi, no en .venv-mail.
"""
import argparse
import json
import os
import sys
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit('falta openpyxl — usar el Python del sistema (3.13), no .venv-mail')


COLUMNAS = [
    ('Rol',              14),
    ('Fecha',            17),
    ('De',               28),
    ('Para',             30),
    ('Copia',            22),
    ('Asunto',           58),
    ('Adjuntos',         42),
    ('Extracto',         70),
    ('Carpeta original', 26),
    ('Archivo (.eml) — abrir con doble click', 62),
]

AZUL = PatternFill('solid', fgColor='1F4E79')
GRIS = PatternFill('solid', fgColor='F2F2F2')
BORDE = Border(*[Side(style='thin', color='BFBFBF')] * 4)


def main():
    ap = argparse.ArgumentParser(description='Excel indice del archivo de correo')
    ap.add_argument('--carpeta', required=True)
    ap.add_argument('--salida', default='_INDICE.xlsx')
    a = ap.parse_args()

    carpeta = a.carpeta
    if not os.path.isdir(carpeta):
        import glob as _g
        cand = _g.glob(carpeta)
        if len(cand) == 1:
            carpeta = cand[0]
        else:
            sys.exit('no existe la carpeta: %s' % a.carpeta)

    jsonl = os.path.join(carpeta, '_indice.jsonl')
    if not os.path.exists(jsonl):
        sys.exit('no encuentro %s — hay que correr _archivoCorreo.py primero' % jsonl)

    filas = []
    with open(jsonl, encoding='utf-8') as fh:
        for linea in fh:
            linea = linea.strip()
            if linea:
                try:
                    filas.append(json.loads(linea))
                except json.JSONDecodeError:
                    continue
    filas.sort(key=lambda d: (d.get('rol', ''), d.get('fecha', '')))
    print('mails en el indice: %d' % len(filas))

    wb = Workbook()
    ws = wb.active
    ws.title = 'ARCHIVO DE CORREO'

    # B2 = primera celda con contenido (convencion de entregables de Barack)
    ws.cell(row=2, column=2, value='ARCHIVO DE CORREO — BARACK ARGENTINA SRL').font = Font(
        bold=True, size=14, color='1F4E79')
    ws.cell(row=3, column=2,
            value='%d mails archivados. Cada fila apunta a un .eml que se abre con doble click, '
                  'sin Outlook ni Exchange.' % len(filas)).font = Font(italic=True, size=9, color='595959')

    ENC = 5
    for i, (titulo, ancho) in enumerate(COLUMNAS):
        c = ws.cell(row=ENC, column=2 + i, value=titulo)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = AZUL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDE
        ws.column_dimensions[get_column_letter(2 + i)].width = ancho

    for j, d in enumerate(filas):
        r = ENC + 1 + j
        valores = [
            d.get('rol', ''),
            d.get('fecha', ''),
            d.get('de', ''),
            d.get('para', ''),
            d.get('cc', ''),
            d.get('asunto', ''),
            ' | '.join(d.get('adjuntos', [])),
            d.get('extracto', ''),
            d.get('carpeta_origen', ''),
            d.get('eml', ''),
        ]
        for i, v in enumerate(valores):
            c = ws.cell(row=r, column=2 + i, value=v)
            c.alignment = Alignment(vertical='top', wrap_text=(i in (5, 6, 7)))
            c.border = BORDE
            if j % 2:
                c.fill = GRIS

    ws.freeze_panes = ws.cell(row=ENC + 1, column=2)
    ws.auto_filter.ref = '%s%d:%s%d' % (get_column_letter(2), ENC,
                                        get_column_letter(1 + len(COLUMNAS)), ENC + len(filas))

    # Hoja 2: el resumen que contesta "que hay aca adentro" sin leer 10.000 filas
    rs = wb.create_sheet('RESUMEN')
    rs.cell(row=2, column=2, value='QUE HAY EN ESTE ARCHIVO').font = Font(bold=True, size=13, color='1F4E79')
    fila = 4
    for titulo, cuenta in (('Mails por rol', Counter(d.get('rol', '') for d in filas)),
                           ('Mails por anio', Counter((d.get('fecha', '') or '????')[:4] for d in filas))):
        rs.cell(row=fila, column=2, value=titulo).font = Font(bold=True)
        fila += 1
        for k, v in sorted(cuenta.items()):
            rs.cell(row=fila, column=2, value=k)
            rs.cell(row=fila, column=3, value=v)
            fila += 1
        fila += 1
    con_adj = sum(1 for d in filas if d.get('adjuntos'))
    rs.cell(row=fila, column=2, value='Mails con adjunto').font = Font(bold=True)
    rs.cell(row=fila, column=3, value=con_adj)
    rs.column_dimensions['B'].width = 52
    rs.column_dimensions['C'].width = 12

    salida = a.salida if os.path.isabs(a.salida) else os.path.join(carpeta, a.salida)
    wb.save(salida)
    print('escrito: %s' % salida)
    print('  filas: %d   hojas: %s' % (len(filas), ', '.join(wb.sheetnames)))


if __name__ == '__main__':
    main()
